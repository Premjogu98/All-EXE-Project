import wx
import wx.lib.scrolledpanel
from Objects import globalvar
import wx.adv
import openpyxl
import wx.html as html


class HTML_Frame(wx.Frame):
    def __init__(self, title, parent=None):
        wx.Frame.__init__(self, parent=parent, title=title,size=(1000, 500), style=wx.DEFAULT_FRAME_STYLE ^ wx.RESIZE_BORDER)
        self.Show()
        self.bSizer = wx.BoxSizer(wx.VERTICAL)
        self.scroll = wx.lib.scrolledpanel.ScrolledPanel(self, -1, size=(1000, 500), pos=(0, 0),
                                                         style=wx.SIMPLE_BORDER)
        self.scroll.SetupScrolling()
        self.scroll.SetBackgroundColour('#FFFFFF')
        self.scroll.SetForegroundColour('Black')
        self.scroll.SetSizer(self.bSizer)

        txt_style = wx.VSCROLL | wx.HSCROLL | wx.TE_READONLY | wx.BORDER_SIMPLE
        self.html = html.HtmlWindow(self.scroll, -1, size=(1000, 500), pos=(0, 0), style=txt_style)
        self.html.SetPage("""
        <h2><font color=\"red\">Error</font><font color=\"black\"> / </font><font color=\"green\"> Completions </font></h2>
        <table align=\"center\" border=\"1\" style=\"width:98%;border-spacing:0;border-collapse: collapse;border:1px solid #666666; margin-top:5px; margin-bottom:5px;\">
          <tr bgcolor=\"#e8eff1\" onmouseover=\"this.style.backgroundColor='#d6edf5'\" onmouseout=\"this.style.backgroundColor=''\">
            <th style="width:25%"><font color=\"#7854E0\">Source Name</font></th>
            <th style="width:100%"><font color=\"red\">Error</font><font color=\"black\"> / </font><font color=\"green\"> Completions </font></th> 
          </tr>"""+str(globalvar.html_string)+"""
          
        </table>
""")


class GUI(wx.Frame):

    def __init__(self,parent,id,title):
        wx.Frame.__init__(self, parent, id, title, size=(1400, 805), style=wx.DEFAULT_FRAME_STYLE ^ wx.RESIZE_BORDER)

        self.bSizer = wx.BoxSizer(wx.VERTICAL)
        self.bSizer1 = wx.BoxSizer(wx.HORIZONTAL)
        self.Top_panel = wx.Panel(self, size=(1400, 75), pos=(0, 0), style=wx.SIMPLE_BORDER)
        self.Top_panel.SetBackgroundColour('#FFFFFF')

        self.scroll = wx.lib.scrolledpanel.ScrolledPanel(self, -1, size=(1370, 670), pos=(7, 82),
                                                         style=wx.SIMPLE_BORDER)
        self.scroll.SetupScrolling()
        self.scroll.SetBackgroundColour('#FFFFFF')
        self.scroll.SetForegroundColour('Black')
        self.scroll.SetSizer(self.bSizer)

        self.Show_Errors = wx.Button(self.Top_panel, -1,label=' Show Error / Completion ', pos=(390, 42))
        self.Show_Errors.Bind(wx.EVT_BUTTON, self.on_new_frame)
        self.Show_Errors.SetBackgroundColour('#7854E0')
        self.Show_Errors.SetForegroundColour('White')

        self.Select_cb = wx.Button(self.Top_panel, label='select All CheckBox', pos=(10, 7))
        self.Select_cb.SetBackgroundColour('#7854E0')
        self.Select_cb.SetForegroundColour('White')
        self.Select_cb.Bind(wx.EVT_BUTTON, self.select_Checkbox)

        self.Unselect_cb = wx.Button(self.Top_panel, label='Unselect All CheckBox', pos=(150, 7))
        self.Unselect_cb.SetBackgroundColour('#7854E0')
        self.Unselect_cb.SetForegroundColour('White')
        self.Unselect_cb.Bind(wx.EVT_BUTTON, self.Unselect_Checkbox)

        self.Number_Of_Source = wx.StaticText(self.Top_panel, label="Number Of Source : ", pos=(12, 47))
        self.Number_Of_Source.SetForegroundColour('Black')

        self.Number_Of_Source_btn = wx.Button(self.Top_panel, -1, label='Get Number Of Selected Source', pos=(180, 42))
        self.Number_Of_Source_btn.Bind(wx.EVT_BUTTON, self.Source_count)
        self.Number_Of_Source_btn.SetBackgroundColour('#7854E0')
        self.Number_Of_Source_btn.SetForegroundColour('White')

        self.Add_source_name = wx.StaticText(self.Top_panel, label="Add Source name", pos=(300, 12))
        self.Add_source_name.SetForegroundColour('Black')
        self.txt_source_name = wx.TextCtrl(self.Top_panel,size = (180,20),pos=(400,10))

        self.Add_source_Link = wx.StaticText(self.Top_panel, label="Add Source Link", pos=(590, 12))
        self.Add_source_Link.SetForegroundColour('Black')
        self.txt_Source_link = wx.TextCtrl(self.Top_panel,size = (180,20),pos=(683,10))

        self.Add_btn = wx.Button(self.Top_panel, label='Add', pos=(883, 7))
        self.Add_btn.SetBackgroundColour('#7854E0')
        self.Add_btn.SetForegroundColour('White')
        self.Add_btn.Bind(wx.EVT_BUTTON, self.Add_Things)

        dpc2 = wx.adv.GenericDatePickerCtrl(self.Top_panel, size=(120, -1),pos=(1010, 10),
                                            style=wx.TAB_TRAVERSAL
                                                  | wx.adv.DP_DROPDOWN

                                                  | wx.adv.DP_ALLOWNONE)
        self.Bind(wx.adv.EVT_DATE_CHANGED, self.Get_Date, dpc2)
        dpc2.SetValue(wx.DateTime.Now())

        self.gobtn = wx.Button(self.Top_panel, label='GO', pos=(1160, 8))
        self.gobtn.SetBackgroundColour('#84FF25')
        self.gobtn.Bind(wx.EVT_BUTTON, self.components)
        self.gobtn.SetForegroundColour('Black')

        self.exitbtn = wx.Button(self.Top_panel, label='EXIT', pos=(1260, 8))
        self.exitbtn.SetBackgroundColour('#FF3838')
        self.exitbtn.Bind(wx.EVT_BUTTON, self.Exit)
        self.exitbtn.SetForegroundColour('White')
        self.exitbtn.SetForegroundColour('Black')

        self.cb_Position_height = 6
        self.TX_Position_height = 20
        self.TX2_Position_height = 12
        self.Panel_Height = 2
        self.Source_count = 1
        self.cb_list = []

        path = "D:\\PycharmProjects\\ALL Exe Project\\source name.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        source_list = []
        number_of_Rows = sheet_obj.max_row
        for i in range(3, number_of_Rows + 1):
            source_name = sheet_obj.cell(row=i, column=1)
            source_Link = sheet_obj.cell(row=i, column=2)
            value = str(source_name.value) + " - " + str(source_Link.value)
            source_list.append(value)

        for source in source_list:
            self.scroll_panel = wx.Panel(self.scroll, size=(1358, 30), pos=(4, self.Panel_Height),
                                         style=wx.SIMPLE_BORDER)
            self.scroll_panel.SetBackgroundColour('#7854E0')
            self.cb = wx.CheckBox(self.scroll_panel, -1, str(source), (12, 6))
            self.cb.SetValue(True)
            self.cb.SetForegroundColour('White')
            self.cb_list.append(self.cb)

            self.Number_Of_Source.SetLabel('Number Of Source Selected : 0')

            self.bSizer.Add(self.scroll_panel, 0, wx.ALL, 5)
            self.scroll.SetupScrolling()

    def Source_count(self,event):
        self.source_count = 0
        for i, self.cb in enumerate(self.cb_list):
            if self.cb.GetValue():
                self.source_count += 1
        self.Number_Of_Source.SetLabel('Number Of Source Selected : '+str(self.source_count)+'')
        self.Number_Of_Source.SetBackgroundColour('green')
        self.Number_Of_Source.SetForegroundColour('Black')

    def on_new_frame(self, event):
        frame = HTML_Frame(title='Errors and Completion')
        # self.frame_number += 1
        # self.Show_Errors.Disable()

    def Add_Things(self,event):
        path = "D:\\PycharmProjects\\ALL Exe Project\\source name.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj1 = wb_obj.active
        source_name = self.txt_source_name.GetValue()
        source_Link = self.txt_Source_link.GetValue()
        if source_name != '' and source_Link != '':
            sheet_obj1.append([source_name, source_Link])
            wb_obj.save(path)
            self.Add_source_name = wx.StaticText(self.Top_panel, label="Add Source name", pos=(300, 12))
            self.Add_source_name.SetForegroundColour('Black')
            self.Add_source_name.SetBackgroundColour('Green')
            self.Add_source_Link = wx.StaticText(self.Top_panel, label="Add Source Link", pos=(590, 12))
            self.Add_source_Link.SetForegroundColour('Black')
            self.Add_source_Link.SetBackgroundColour('Green')
            self.txt_source_name.Clear()
            self.txt_Source_link.Clear()

        else:
            self.Add_source_name = wx.StaticText(self.Top_panel, label="Insert Source Name", pos=(295, 12))
            self.Add_source_name.SetForegroundColour('White')
            self.Add_source_name.SetBackgroundColour('Red')
            print('Insert Source Name On Text Box')
            self.Add_source_Link = wx.StaticText(self.Top_panel, label="Insert Source Link", pos=(585, 12))
            self.Add_source_Link.SetForegroundColour('White')
            self.Add_source_Link.SetBackgroundColour('Red')
            print('Insert Source Link On Text Box')

    def components(self,event):
        source_name_list = []
        for i, self.cb in enumerate(self.cb_list):
            if self.cb.GetValue():
                Source_name = self.cb.GetLabelText()
                Source_name = Source_name.partition("-")[0].strip()
                source_name_list.append(Source_name)
        count = 0
        for source in source_name_list:
            from Source_connection.All_source import Match_source
            Match_source(source)
            count += 1
            print('\n Source Name '+str(count)+' : '+str(source) + '  ----> Done')
            print('===================================================================\n')
        wx.MessageBox('All Process Are Done', 'All Exe Project', wx.OK | wx.ICON_INFORMATION)

    def Exit(self, event):
        dlg = wx.MessageDialog(None, "Kya Aap Ko yaha Se Prasthan (EXIT) karna Hai !!!!", 'Updater', wx.YES_NO | wx.ICON_WARNING)
        result = dlg.ShowModal()
        if result == wx.ID_YES:
            self.Destroy()
        else:
            pass

    def Get_Date(self, evt):
        sel_date = evt.GetDate()
        globalvar.From_Date = sel_date.Format("%Y-%m-%d")
        print(f'Selected Date {globalvar.From_Date}')

    def Unselect_Checkbox(self,event):
        for Checkbox in self.cb_list:
            Checkbox.SetValue(False)

    def select_Checkbox(self,event):
        for Checkbox in self.cb_list:
            Checkbox.SetValue(True)


app = wx.App()
frame = GUI(parent=None, id=-1, title="All Exe Project")
frame.Show()
app.MainLoop()
